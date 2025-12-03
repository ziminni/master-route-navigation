from PyQt6 import QtWidgets, QtGui, QtCore
from PyQt6.QtCore import QTimer, Qt, QAbstractTableModel
from PyQt6.QtGui import QColor
from PyQt6.QtWidgets import QMessageBox, QDialog, QVBoxLayout, QCalendarWidget, QFileDialog, QTabWidget, QGridLayout, QPushButton, QLabel, QScrollArea, QHBoxLayout
import os
import json
import datetime

from reportlab.platypus import SimpleDocTemplate, Paragraph
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import letter
import openpyxl
from openpyxl.styles import Font

from typing import Dict, List
from ..Base.manager_base import ManagerBase
from ..Base.faculty_admin_base import FacultyAdminBase
from widgets.orgs_custom_widgets.dialogs import CreateOrgDialog, ArchiveConfirmDialog
from widgets.orgs_custom_widgets.cards import CollegeOrgCard, ArchivedOrgCard
from ui.Organization.audit_logs_ui import Ui_audit_logs_widget
from ui.Organization.generate_reports_ui import Ui_generate_reports_widget

class AuditLogModel(QAbstractTableModel):
    """Model for displaying audit logs."""
    def __init__(self, data: List[Dict], parent=None):
        super().__init__(parent)
        self._data = data
        self._headers = ["Timestamp", "Actor", "Action", "Organization", "Subject", "Details"]

    def rowCount(self, parent=QtCore.QModelIndex()) -> int:
        return len(self._data)

    def columnCount(self, parent=QtCore.QModelIndex()) -> int:
        return len(self._headers)

    def data(self, index, role=Qt.ItemDataRole.DisplayRole):
        if not index.isValid():
            return None
        
        try:
            row_data = self._data[index.row()]
            col_name = self._headers[index.column()]
        except IndexError:
            return None
        
        if role == Qt.ItemDataRole.DisplayRole:
            value = row_data.get(col_name.lower())
            if col_name == "Timestamp":
                try:
                    dt = datetime.datetime.fromisoformat(value)
                    return dt.strftime("%Y-%m-%d %I:%M:%S %p")
                except (ValueError, TypeError):
                    return value
            return value
        
        # Consistent styling: use alternating colors (as in organization_view_base._apply_table_style)
        if role == Qt.ItemDataRole.BackgroundRole:
            return QColor("#f6f8fa") if index.row() % 2 == 1 else QColor("white")

        return None

    def headerData(self, section, orientation, role=Qt.ItemDataRole.DisplayRole):
        if role == Qt.ItemDataRole.DisplayRole and orientation == Qt.Orientation.Horizontal:
            return self._headers[section]
        return None

class Admin(ManagerBase, FacultyAdminBase):
    """
    Admin view. Inherits from ManagerBase (for management methods) and
    FacultyAdminBase (for the non-student UI layout).
    Adds admin-specific features like the admin menu, create button,
    audit logs, and report generation pages.
    
    MRO: Admin -> ManagerBase -> FacultyAdminBase -> OrganizationViewBase -> User
    """
    
    def __init__(self, admin_name: str):
        FacultyAdminBase.__init__(self, name=admin_name)
        ManagerBase.__init__(self)
        
        # Set admin role for API calls
        self.user_role = "admin"
        
        # Initialize members_dict for storing full member data
        self.members_dict = {}
        
        self.report_start_date: datetime.date | None = None
        self.report_end_date: datetime.date | None = None

        self.audit_no_logs_label: QtWidgets.QLabel | None = None 
        self.audit_prev_btn: QtWidgets.QPushButton | None = None
        self.audit_next_btn: QtWidgets.QPushButton | None = None
        self.audit_page_label: QtWidgets.QLabel | None = None
        self.current_audit_page = 0
        self.audit_items_per_page = 100 # New limit
        self.cached_audit_logs: List[Dict] = []
        self.current_audit_search: str = ""
        
        self.archive_btn = None
        self.archived_widget = None
        self.archived_ui = None
        self.archived_tab_widget = None
        self.archived_index = None
        self.from_archived = False
        self.current_archived_page = 0
        self.items_per_page = 10
        self.archived_search_text = ""
        
        self.org_prev_btn = None
        self.org_next_btn = None
        self.org_page_label = None
        self.branch_prev_btn = None
        self.branch_next_btn = None
        self.branch_page_label = None
        
        self.last_report_logs: List[Dict] = []
        self.last_report_data: List[Dict] = []
        self.last_report_org: str = ""
        self.last_report_type: str = ""
        self.last_report_header_text: str = ""
        
        self._setup_admin_menu()
        self._setup_create_button()
        
        self._setup_audit_logs_page()
        self._setup_generate_reports_page()
        
        self.ui.stacked_widget.currentChanged.connect(self._on_stack_changed)
        
        self.load_orgs()

    # --- ADDED: Cooldown Bypass Overrides ---
    
    def edit_member(self, row: int) -> None:
        """Admin override to bypass manager cooldown."""
        super().edit_member(row, bypass_cooldown=True)

    def kick_member(self, row: int) -> None:
        """Admin override to bypass manager cooldown."""
        super().kick_member(row, bypass_cooldown=True)

    def update_officer_in_org(self, updated_officer: Dict) -> None:
        """Admin override to bypass manager cooldown."""
        super().update_officer_in_org(updated_officer, bypass_cooldown=True)

    # --- End Cooldown Bypass ---

    def showEvent(self, event):
        """Override showEvent to ensure proper initial positioning after the widget is shown."""
        super().showEvent(event)
        QTimer.singleShot(0, self._reposition_create_button)
    
    def _on_stack_changed(self, index: int) -> None:
        """Handle stacked widget changes to reposition floating buttons."""
        if index == 0:
            QTimer.singleShot(0, self._reposition_create_button)
    
    def _setup_admin_menu(self) -> None:
        """Set up the admin menu button with dropdown options beside the search bar."""
        self.menu_btn = QtWidgets.QPushButton(self.ui.landing_page)
        self.menu_btn.setStyleSheet("background-color: transparent;")
        self.menu_btn.setObjectName("menu_btn")
        self.menu_btn.setCursor(QtGui.QCursor(QtCore.Qt.CursorShape.PointingHandCursor))
        self.menu_btn.setText("")
        
        icon = QtGui.QIcon()
        
        menu_icon_path = os.path.join(
            os.path.dirname(__file__), 
            "..", "..", "..", "..",
            "assets", "organization", "icons", "menu.png"
        )

        icon.addPixmap(
            QtGui.QPixmap(menu_icon_path), 
            QtGui.QIcon.Mode.Normal, 
            QtGui.QIcon.State.Off
        )
        self.menu_btn.setIcon(icon)
        self.menu_btn.setIconSize(QtCore.QSize(30, 30))
        
        self.ui.horizontalLayout_2.addWidget(self.menu_btn)
        
        self.admin_menu = QtWidgets.QMenu(self)
        
        self.admin_menu.setStyleSheet("""
            QMenu {
                background-color: white;
                border: 1px solid #ccc;
                border-radius: 5px;
            }
            QMenu::item {
                padding: 5px 25px 5px 25px; 
                color: black;
                border-radius: 5px;
            }
            QMenu::item:selected {
                /* This is the hover/selection effect */
                background-color: #FDC601; /* Your project's yellow */
                color: #084924;        /* Your project's dark green */
            }
            QMenu::separator {
                height: 1px;
                background: #e0e0e0;
                margin: 5px 0px 5px 0px;
            }
        """)

        audit_logs_action = self.admin_menu.addAction("Audit Logs")
        audit_logs_action.triggered.connect(self._open_audit_logs)
        
        archive_action = self.admin_menu.addAction("View Archived")
        archive_action.triggered.connect(self._show_archived_view)
        
        self.menu_btn.clicked.connect(self._show_admin_menu)
    
    def _show_admin_menu(self) -> None:
        """Display the admin dropdown menu."""
        button_pos = self.menu_btn.mapToGlobal(QtCore.QPoint(0, self.menu_btn.height()))
        self.admin_menu.exec(button_pos)
    
    def _setup_create_button(self) -> None:
        """Set up the create organization/branch button in the lower right."""
        self.create_btn = QtWidgets.QPushButton("+ Create Organization/Branch", self.ui.landing_page)
        self.create_btn.setStyleSheet("""
            QPushButton {
                background-color: #FDC601;
                color: white;
                border-radius: 10px;
                padding: 15px 20px;
                font-size: 14px;
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #084924;
            }
        """)
        self.create_btn.setCursor(QtGui.QCursor(QtCore.Qt.CursorShape.PointingHandCursor))
        self.create_btn.clicked.connect(self._create_organization)
        
        self._reposition_create_button()
        self.create_btn.raise_()

    def _reposition_create_button(self) -> None:
        """Helper to set the geometry of the create button based on landing_page size."""
        if hasattr(self, 'create_btn'):
            BUTTON_WIDTH = 260
            BUTTON_HEIGHT = 50
            MARGIN_RIGHT = 30
            MARGIN_BOTTOM = 70

            self.create_btn.setGeometry(
                self.ui.landing_page.width() - BUTTON_WIDTH - MARGIN_RIGHT,
                self.ui.landing_page.height() - BUTTON_HEIGHT - MARGIN_BOTTOM,
                BUTTON_WIDTH,
                BUTTON_HEIGHT
            )
            self.create_btn.raise_()

    def resizeEvent(self, event) -> None:
        """Handle resize events to reposition the create button."""
        super().resizeEvent(event)
        if self.ui.stacked_widget.currentIndex() == 0:
            self._reposition_create_button()
    
    def _setup_audit_logs_page(self) -> None:
        """Load the audit logs UI as a new stacked widget page and add pagination controls."""
        self.audit_logs_page = QtWidgets.QWidget()
        self.audit_logs_ui = Ui_audit_logs_widget()
        self.audit_logs_ui.setupUi(self.audit_logs_page)
        
        self.audit_no_logs_label = QtWidgets.QLabel("No logs.", self.audit_logs_ui.audit_table_container)
        self.audit_no_logs_label.setAlignment(QtCore.Qt.AlignmentFlag.AlignCenter)
        self.audit_no_logs_label.setStyleSheet("font-size: 20px; color: #888;")
        sizePolicy = QtWidgets.QSizePolicy(
            QtWidgets.QSizePolicy.Policy.Preferred,
            QtWidgets.QSizePolicy.Policy.Expanding
        )
        self.audit_no_logs_label.setSizePolicy(sizePolicy)
        self.audit_logs_ui.verticalLayout_17.addWidget(self.audit_no_logs_label)
        self.audit_no_logs_label.hide()
        
        self.audit_logs_ui.audit_back_btn.clicked.connect(
            lambda: self.ui.stacked_widget.setCurrentIndex(0)
        )
        
        self.audit_logs_ui.audit_search_btn.clicked.connect(self._perform_audit_search)
        self.audit_logs_ui.audit_line_edit.returnPressed.connect(self._perform_audit_search)
        
        # --- Audit Logs Pagination Controls ---
        pagination_layout = QHBoxLayout()
        self.audit_prev_btn = QPushButton("Previous")
        self.audit_prev_btn.setStyleSheet("""
            QPushButton { background-color: #084924; color: white; border-radius: 5px; padding: 5px 15px; font-weight: bold;}
            QPushButton:hover { background-color: #FDC601; color: #084924;}
        """)
        self.audit_prev_btn.clicked.connect(lambda: self._change_audit_page(-1))
        
        self.audit_next_btn = QPushButton("Next")
        self.audit_next_btn.setStyleSheet("""
            QPushButton { background-color: #084924; color: white; border-radius: 5px; padding: 5px 15px; font-weight: bold;}
            QPushButton:hover { background-color: #FDC601; color: #084924;}
        """)
        self.audit_next_btn.clicked.connect(lambda: self._change_audit_page(1))
        
        self.audit_page_label = QLabel("Page 1 of 1")
        
        pagination_layout.addWidget(self.audit_prev_btn)
        pagination_layout.addWidget(self.audit_page_label)
        pagination_layout.addWidget(self.audit_next_btn)
        
        self.audit_logs_ui.verticalLayout_17.addLayout(pagination_layout)
        # --- End Pagination Controls ---
        
        self.ui.stacked_widget.addWidget(self.audit_logs_page)
    
    def _setup_generate_reports_page(self) -> None:
        """Load the generate reports UI as a new stacked widget page."""
        self.reports_page = QtWidgets.QWidget()
        self.reports_ui = Ui_generate_reports_widget()
        self.reports_ui.setupUi(self.reports_page)
        
        self.reports_ui.report_preview_text = QtWidgets.QPlainTextEdit(self.reports_ui.preview_frame)
        self.reports_ui.report_preview_text.setReadOnly(True)
        self.reports_ui.report_preview_text.setStyleSheet("background-color: transparent; border: none; font-family: 'Courier New', monospace;")
        
        preview_layout = QtWidgets.QVBoxLayout(self.reports_ui.preview_frame)
        preview_layout.setContentsMargins(10, 10, 10, 10)
        preview_layout.addWidget(self.reports_ui.report_preview_text)
        
        self.reports_ui.back_btn_reports.clicked.connect(
            lambda: self.ui.stacked_widget.setCurrentIndex(0)
        )
        
        self.reports_ui.generate_report_btn.clicked.connect(self.generate_report)
        self.reports_ui.pushButton.clicked.connect(self._select_start_date)
        self.reports_ui.pushButton_2.clicked.connect(self._select_end_date)
        self.reports_ui.download_pdf_btn.clicked.connect(self._download_pdf)
        self.reports_ui.download_excel_btn.clicked.connect(self._download_excel)
        
        self.ui.stacked_widget.addWidget(self.reports_page)
        
    def _generate_reports(self) -> None:
        """Navigate to generate reports page."""
        try:
            self._populate_reports_orgs() 
            self.reports_ui.report_preview_text.clear()
            self.report_start_date = None
            self.report_end_date = None
            self.reports_ui.pushButton.setText("Start Date")
            self.reports_ui.pushButton_2.setText("End Date")
            
            self.last_report_logs = []
            self.last_report_data = []
            self.last_report_org = ""
            self.last_report_type = ""
            self.last_report_header_text = ""
            
            self.ui.stacked_widget.setCurrentIndex(4)
        except Exception as e:
            import traceback
            traceback.print_exc()
            QMessageBox.critical(self, "Error", f"Error opening Generate Reports: {str(e)}")

    def _open_audit_logs(self) -> None:
        """Navigate to audit logs page and initialize data."""
        self.current_audit_page = 0
        self.current_audit_search = ""
        self.audit_logs_ui.audit_line_edit.clear()
        self._load_all_audit_logs_into_cache() # Load all logs once
        self.load_audit_logs_data() # Display first page
        self.ui.stacked_widget.setCurrentIndex(3)
        
    def _load_all_audit_logs_into_cache(self) -> None:
        """Loads all audit logs from database API and caches them."""
        logs = []
        try:
            # Fetch logs from database API
            from services.organization_api_service import OrganizationAPIService
            response = OrganizationAPIService.get_logs(limit=1000)
            
            if response.get('success'):
                api_logs = response.get('data', [])
                # Transform API log format to match expected format for display
                for log in api_logs:
                    logs.append({
                        'timestamp': log.get('date_created', ''),
                        'actor': log.get('user_name', f"User #{log.get('user_id', 'Unknown')}"),
                        'action': log.get('action_display', log.get('action', '')),
                        'organization': log.get('target_type', ''),
                        'subject': log.get('target_name', f"{log.get('target_type', '')} #{log.get('target_id', '')}"),
                        'details': f"{log.get('action_display', log.get('action', ''))} on {log.get('target_name', log.get('target_type', ''))}"
                    })
            else:
                pass  # API error, logs will remain empty
        except Exception:
            logs = []
            
        self.cached_audit_logs = logs

    def _show_archived_view(self) -> None:
        if self.archived_widget:
            self.ui.stacked_widget.removeWidget(self.archived_widget)
            self.archived_widget.deleteLater()
        
        self.archived_widget = QtWidgets.QWidget()
        self.archived_ui = Ui_audit_logs_widget()
        self.archived_ui.setupUi(self.archived_widget)
        
        self.archived_ui.audit_logs_label.setText("Archive")
        self.archived_ui.table_header_label.setText("")
        self.archived_ui.adult_hr_line.hide()
        self.archived_ui.audit_line_edit.setPlaceholderText("Search archived...")
        self.archived_ui.audit_burger_ic.hide()
        
        self.archived_ui.audit_back_btn.clicked.connect(lambda: self.ui.stacked_widget.setCurrentIndex(0))
        self.archived_ui.audit_search_btn.clicked.connect(self._perform_archived_search)
        self.archived_ui.audit_line_edit.returnPressed.connect(self._perform_archived_search)
        
        # Remove table
        self.archived_ui.verticalLayout_17.removeWidget(self.archived_ui.audit_table_view)
        self.archived_ui.audit_table_view.deleteLater()
        
        # Add tabs
        self.archived_tab_widget = QTabWidget()
        org_tab = QtWidgets.QWidget()
        branch_tab = QtWidgets.QWidget()
        
        self.archived_tab_widget.addTab(org_tab, "Organizations")
        self.archived_tab_widget.addTab(branch_tab, "Branches")
        
        # Org tab layout
        org_layout = QVBoxLayout(org_tab)
        self.archived_org_grid = QGridLayout()
        org_scroll = QScrollArea()
        org_scroll.setWidgetResizable(True)
        org_widget = QtWidgets.QWidget()
        org_widget.setLayout(self.archived_org_grid)
        org_scroll.setWidget(org_widget)
        org_layout.addWidget(org_scroll)
        
        # Branch tab layout
        branch_layout = QVBoxLayout(branch_tab)
        self.archived_branch_grid = QGridLayout()
        branch_scroll = QScrollArea()
        branch_scroll.setWidgetResizable(True)
        branch_widget = QtWidgets.QWidget()
        branch_widget.setLayout(self.archived_branch_grid)
        branch_scroll.setWidget(branch_widget)
        branch_layout.addWidget(branch_scroll)
        
        # Pagination for both
        self._add_pagination_controls(org_layout, is_branch=False)
        self._add_pagination_controls(branch_layout, is_branch=True)
        
        self.archived_ui.verticalLayout_17.addWidget(self.archived_tab_widget)
        
        self.archived_tab_widget.currentChanged.connect(lambda idx: self._load_archived_page(idx == 1))
        
        self.archived_index = self.ui.stacked_widget.addWidget(self.archived_widget)
        self.ui.stacked_widget.setCurrentWidget(self.archived_widget)
        
        # Load on show
        self._load_archived_page(is_branch=False)
        self._load_archived_page(is_branch=True)

    def _perform_archived_search(self):
        self.archived_search_text = self.archived_ui.audit_line_edit.text().strip().lower()
        self.current_archived_page = 0
        is_branch = self.archived_tab_widget.currentIndex() == 1
        self._load_archived_page(is_branch)

    def _add_pagination_controls(self, layout: QVBoxLayout, is_branch: bool):
        # CONSISTENCY FIX: Apply style to pagination buttons
        PAGINATION_BTN_STYLE = """
            QPushButton {
                background-color: #084924; /* Dark Green */
                color: white;
                border-radius: 5px;
                padding: 5px 15px; 
                font-weight: bold;
            }
            QPushButton:hover {
                background-color: #FDC601; /* Yellow */
                color: #084924;
            }
        """

        pagination_layout = QHBoxLayout()
        prev_btn = QPushButton("Previous")
        prev_btn.setStyleSheet(PAGINATION_BTN_STYLE)
        prev_btn.clicked.connect(lambda: self._change_archived_page(-1, is_branch))
        next_btn = QPushButton("Next")
        next_btn.setStyleSheet(PAGINATION_BTN_STYLE)
        next_btn.clicked.connect(lambda: self._change_archived_page(1, is_branch))
        page_label = QLabel("Page 1")
        pagination_layout.addWidget(prev_btn)
        pagination_layout.addWidget(page_label)
        pagination_layout.addWidget(next_btn)
        layout.addLayout(pagination_layout)
        
        if not is_branch:
            self.org_prev_btn = prev_btn
            self.org_next_btn = next_btn
            self.org_page_label = page_label
        else:
            self.branch_prev_btn = prev_btn
            self.branch_next_btn = next_btn
            self.branch_page_label = page_label

    def _change_archived_page(self, delta: int, is_branch: bool):
        archived = self.get_archived(is_branch=is_branch, search_text=self.archived_search_text)
        total_pages = (len(archived) + self.items_per_page - 1) // self.items_per_page
        new_page = max(0, min(self.current_archived_page + delta, total_pages - 1))
        if new_page != self.current_archived_page:
            self.current_archived_page = new_page
            self._load_archived_page(is_branch)

    def _load_archived_page(self, is_branch: bool):
        archived = self.get_archived(is_branch=is_branch, search_text=self.archived_search_text)
        grid = self.archived_org_grid if not is_branch else self.archived_branch_grid
        self._clear_grid(grid)
        
        start = self.current_archived_page * self.items_per_page
        end = start + self.items_per_page
        for i, item in enumerate(archived[start:end]):
            card = ArchivedOrgCard(self._get_logo_path(item["logo_path"]), item["name"], item, self)
            col = i % 5
            row = i // 5
            grid.addWidget(card, row, col, alignment=QtCore.Qt.AlignmentFlag.AlignTop | QtCore.Qt.AlignmentFlag.AlignHCenter)
        
        if len(archived) == 0:
            self._add_no_record_label(grid)
        
        total_pages = (len(archived) + self.items_per_page - 1) // self.items_per_page
        prev_btn = self.org_prev_btn if not is_branch else self.branch_prev_btn
        next_btn = self.org_next_btn if not is_branch else self.branch_next_btn
        page_label = self.org_page_label if not is_branch else self.branch_page_label
        
        prev_btn.setEnabled(self.current_archived_page > 0)
        next_btn.setEnabled(self.current_archived_page < total_pages - 1)
        page_label.setText(f"Page {self.current_archived_page + 1} of {max(1, total_pages)}")
    
    def get_archived(self, is_branch: bool = None, search_text: str = "") -> List[Dict]:
        """Fetch archived organizations from the API."""
        from services.organization_api_service import OrganizationAPIService
        
        api_response = OrganizationAPIService.fetch_organizations(
            search_query=search_text if search_text else None,
            role='admin',
            include_archived=True
        )
        
        archived = []
        if api_response.get('success'):
            organizations_data = api_response.get('data', [])
            for org in organizations_data:
                org_dict = {
                    "id": org.get('id'),
                    "name": org.get('name'),
                    "description": org.get('description', ''),
                    "objectives": org.get('objectives', ''),
                    "status": org.get('status', 'active'),
                    "logo_path": org.get('logo_path', 'No Photo'),
                    "org_level": org.get('org_level', 'col'),
                    "created_at": org.get('created_at'),
                    "is_branch": False,  # TODO: Check main_org field
                    "is_archived": org.get('is_archived', False),
                    "is_active": org.get('is_active', True),
                }
                # Filter by is_branch if specified
                if is_branch is None or org_dict["is_branch"] == is_branch:
                    archived.append(org_dict)
        
        return archived

    def _create_organization(self) -> None:
        """Handle create organization action."""
        dialog = CreateOrgDialog(self)
        dialog.exec()
        self.load_orgs()
            
    def _return_to_prev_page(self) -> None:
        """Navigate back, handling admin pages."""
        current_index = self.ui.stacked_widget.currentIndex()
        
        if current_index in [3, 4] or current_index == self.archived_index:
            self.ui.stacked_widget.setCurrentIndex(0)
            self.from_archived = False
        elif current_index == 1 and self.from_archived:
            self.ui.stacked_widget.setCurrentIndex(self.archived_index)
            self.from_archived = False
        else:
            super()._return_to_prev_page()

    def _change_audit_page(self, delta: int) -> None:
        """Changes the current audit log page."""
        # Calculate total pages based on currently filtered/cached logs
        total_logs = len(self.cached_audit_logs)
        if self.current_audit_search:
            # Re-filter if search text is active
            filtered_logs = [
                log for log in self.cached_audit_logs
                if any(self.current_audit_search in str(val).lower() for val in log.values() if val)
            ]
            total_logs = len(filtered_logs)
        
        total_pages = (total_logs + self.audit_items_per_page - 1) // self.audit_items_per_page
        
        new_page = max(0, min(self.current_audit_page + delta, total_pages - 1))
        if new_page != self.current_audit_page:
            self.current_audit_page = new_page
            self.load_audit_logs_data(self.current_audit_search, use_cache=True)
            
    def load_audit_logs_data(self, search_text: str = "", use_cache: bool = False) -> None:
        """Load and filter audit logs into the table, using pagination."""
        
        if not use_cache:
            self._load_all_audit_logs_into_cache()
            self.current_audit_search = search_text.strip().lower()
            self.current_audit_page = 0
            
        logs = self.cached_audit_logs
        
        filtered_logs = []
        if self.current_audit_search:
            for log in logs:
                if any(self.current_audit_search in str(val).lower() for val in log.values() if val):
                    filtered_logs.append(log)
        else:
            filtered_logs = logs
        
        total_logs = len(filtered_logs)
        total_pages = (total_logs + self.audit_items_per_page - 1) // self.audit_items_per_page
        
        start_index = self.current_audit_page * self.audit_items_per_page
        end_index = start_index + self.audit_items_per_page
        paginated_logs = filtered_logs[start_index:end_index]
        
        model = AuditLogModel(paginated_logs)
        self.audit_logs_ui.audit_table_view.setModel(model)
        
        # Apply style consistent with organization_view_base._apply_table_style
        self.audit_logs_ui.audit_table_view.setAlternatingRowColors(True)
        self.audit_logs_ui.audit_table_view.horizontalHeader().setSectionResizeMode(QtWidgets.QHeaderView.ResizeMode.ResizeToContents)
        self.audit_logs_ui.audit_table_view.horizontalHeader().setSectionResizeMode(5, QtWidgets.QHeaderView.ResizeMode.Stretch)
        self.audit_logs_ui.audit_table_view.verticalHeader().setVisible(False)
        self.audit_logs_ui.audit_table_view.setStyleSheet("QTableView { background-color: white; border: none; }")

        # Update Pagination Controls
        self.audit_prev_btn.setEnabled(self.current_audit_page > 0)
        self.audit_next_btn.setEnabled(self.current_audit_page < total_pages - 1)
        self.audit_page_label.setText(f"Page {self.current_audit_page + 1} of {max(1, total_pages)}")

        if not paginated_logs:
            self.audit_logs_ui.audit_table_view.hide()
            if self.audit_no_logs_label:
                self.audit_no_logs_label.show()
        else:
            self.audit_logs_ui.audit_table_view.show()
            if self.audit_no_logs_label:
                self.audit_no_logs_label.hide()

    def _perform_audit_search(self) -> None:
        """Handle audit logs search."""
        search_text = self.audit_logs_ui.audit_line_edit.text().strip().lower()
        self.load_audit_logs_data(search_text, use_cache=False)

    def _populate_reports_orgs(self) -> None:
        """Populate organizations in the reports dropdown."""
        orgs = self._load_data()
        self.reports_ui.select_org_dd.clear()
        for org in orgs:
            if not org.get("is_branch", False):
                self.reports_ui.select_org_dd.addItem(org.get("name", "Unknown"))
                for branch in org.get("branches", []):
                    self.reports_ui.select_org_dd.addItem(f"  - {branch.get('name', 'Unknown')}")

    def _create_calendar_dialog(self, current_date=None) -> tuple[QDialog, QCalendarWidget]:
        """Helper to create a modal calendar dialog."""
        dialog = QDialog(self)
        dialog.setWindowTitle("Select Date")
        layout = QVBoxLayout(dialog)
        calendar = QCalendarWidget(dialog)
        if current_date:
            calendar.setSelectedDate(QtCore.QDate(current_date))
        layout.addWidget(calendar)
        dialog.setLayout(layout)
        return dialog, calendar

    def _select_start_date(self) -> None:
        """Open a calendar to select the report start date."""
        dialog, calendar = self._create_calendar_dialog(self.report_start_date)
        
        def on_date_clicked(date):
            self.report_start_date = date.toPyDate()
            self.reports_ui.pushButton.setText(f"Start: {self.report_start_date.strftime('%Y-%m-%d')}")
            dialog.accept()

        calendar.clicked.connect(on_date_clicked)
        dialog.exec()

    def _select_end_date(self) -> None:
        """Open a calendar to select the report end date."""
        dialog, calendar = self._create_calendar_dialog(self.report_end_date)
        
        def on_date_clicked(date):
            self.report_end_date = date.toPyDate()
            self.reports_ui.pushButton_2.setText(f"End: {self.report_end_date.strftime('%Y-%m-%d')}")
            dialog.accept()

        calendar.clicked.connect(on_date_clicked)
        dialog.exec()

    def _generate_event_history_report(self, logs: List[Dict]) -> str:
        """Generates the body for the 'Event History' report."""
        if not logs:
            return "No activities found in this period."
        
        report_text = ""
        for log in sorted(logs, key=lambda x: x['timestamp']):
            dt = datetime.datetime.fromisoformat(log['timestamp']).strftime('%Y-%m-%d %I:%M %p')
            report_text += f"[{dt}] {log['actor']} performed {log['action']}.\\n"
            report_text += f"    -> Subject: {log.get('subject', 'N/A')}\\n"
            report_text += f"    -> Details: {log.get('details', 'N/A')}\\n\\n"
        return report_text

    def _generate_membership_report(self, logs: List[Dict]) -> str:
        """Generates the body for the 'Membership' report."""
        report_logs = [
            log for log in logs 
            if log["action"] in ["ACCEPT_APPLICANT", "KICK_MEMBER", "EDIT_MEMBER"]
        ]
        if not report_logs:
            return "No membership changes found in this period."
        
        report_text = ""
        for log in sorted(report_logs, key=lambda x: x['timestamp']):
            dt = datetime.datetime.fromisoformat(log['timestamp']).strftime('%Y-%m-%d %I:%M %p')
            report_text += f"[{dt}] {log['actor']} performed {log['action']} on {log['subject']}. \\n    Details: {log['details']}\\n"
        return report_text

    def _generate_officer_list_report(self, org_name: str) -> str:
        """
        Generates the body for the 'Officer List' report.
        Also caches the officer data in self.last_report_data.
        """
        report_text = "This report shows the CURRENT officer list and is not affected by the date range.\\n\\n"
        all_orgs = self._load_data()
        found_org = None
        for org in all_orgs:
            if org['name'] == org_name:
                found_org = org
                break
            for branch in org.get("branches", []):
                if branch['name'] == org_name:
                    found_org = branch
                    break
            if found_org:
                break
        
        if not found_org or not found_org.get("officers"):
            self.last_report_data = []
            return "No officers found for this organization."
        
        self.last_report_data = found_org.get("officers", [])
        for officer in self.last_report_data:
            report_text += f" - {officer.get('name', 'N/A'):<30} | {officer.get('position', 'N/A')}\\n"
        return report_text

    def _generate_summary_report(self, logs: List[Dict]) -> str:
        """Generates the body for the 'Summary' report."""
        if not logs:
            return "No activities found in this period."
        
        actions = {}
        for log in logs:
            action_type = log.get('action', 'UNKNOWN')
            actions[action_type] = actions.get(action_type, 0) + 1
        
        report_text = "Summary of all actions in this period:\\n\\n"
        for action, count in sorted(actions.items()):
            report_text += f" - {action:<25} | {count} time(s)\\n"
        return report_text

    # --- Main Report Generation Method ---
    
    def generate_report(self) -> None:
        """Generate report based on selections and display in preview."""
        org_name = self.reports_ui.select_org_dd.currentText().strip().lstrip("- ")
        report_type = self.reports_ui.report_type_dd.currentText()
        
        if org_name == "Select Organization":
            QMessageBox.warning(self, "Select Organization", "Please select an organization or branch.")
            return
            
        start_date = self.report_start_date
        end_date = self.report_end_date
        
        if start_date and end_date and start_date > end_date:
            QMessageBox.warning(self, "Invalid Date Range", "Start date must be before end date.")
            return
            
        date_range_str = "for all time"
        if start_date and end_date:
            date_range_str = f"from {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}"
        elif start_date:
            date_range_str = f"from {start_date.strftime('%Y-%m-%d')} onwards"
        elif end_date:
            date_range_str = f"up to {end_date.strftime('%Y-%m-%d')}"
        
        logs = []
        try:
            # Fetch logs from database API
            from services.organization_api_service import OrganizationAPIService
            response = OrganizationAPIService.get_logs(limit=1000)
            
            if response.get('success'):
                api_logs = response.get('data', [])
                # Transform API log format to match expected format for display
                for log in api_logs:
                    logs.append({
                        'timestamp': log.get('date_created', ''),
                        'actor': log.get('user_name', f"User #{log.get('user_id', 'Unknown')}"),
                        'action': log.get('action_display', log.get('action', '')),
                        'organization': log.get('target_name', log.get('target_type', '')),
                        'subject': log.get('target_name', f"{log.get('target_type', '')} #{log.get('target_id', '')}"),
                        'details': f"{log.get('action_display', log.get('action', ''))} on {log.get('target_name', log.get('target_type', ''))}"
                    })
        except Exception as e:
            self.reports_ui.report_preview_text.setPlainText(f"Error loading audit logs: {e}")
            return

        # For now, show all logs since we don't have org-specific filtering yet
        # org_logs = [log for log in logs if log.get("organization") == org_name]
        org_logs = logs  # Show all logs
        
        date_filtered_logs = []
        for log in org_logs:
            try:
                log_dt = datetime.datetime.fromisoformat(log['timestamp']).date()
                if start_date and log_dt < start_date:
                    continue
                if end_date and log_dt > end_date:
                    continue
                date_filtered_logs.append(log)
            except (ValueError, TypeError):
                continue
        
        report_text = f"Report Type: {report_type}\\n"
        report_text += f"Organization/Branch: {org_name}\\n"
        report_text += f"Date Range: {date_range_str}\\n"
        report_text += f"Generated by: {self.name}\\n"
        report_text += f"Generated on: {datetime.datetime.now().strftime('%Y-%m-%d %I:%M %p')}\\n"
        report_text += "=" * 50 + "\\n\\n"
        
        self.last_report_header_text = "\\n".join(report_text.split('\\n')[:6])
        
        body_text = ""
        if report_type == "Event History (Default)":
            body_text = self._generate_event_history_report(date_filtered_logs)
        elif report_type == "Membership":
            body_text = self._generate_membership_report(date_filtered_logs)
        elif report_type == "Officer List":
            body_text = self._generate_officer_list_report(org_name)
        elif report_type == "Summary":
            body_text = self._generate_summary_report(date_filtered_logs)
        else:
            body_text = f"Report type '{report_type}' is not yet implemented."
        
        self.reports_ui.report_preview_text.setPlainText(report_text + body_text)
        
        self.last_report_logs = date_filtered_logs
        self.last_report_org = org_name
        self.last_report_type = report_type

    def _download_pdf(self) -> None:
        """Saves the content of the report preview as a PDF."""
        report_text = self.reports_ui.report_preview_text.toPlainText()
        if not report_text:
            QMessageBox.warning(self, "Empty Report", "Cannot download an empty report. Please generate a report first.")
            return

        org_name = self.last_report_org
        report_type = self.last_report_type
        suggested_name = f"{org_name}_{report_type}_Report.pdf".replace(" ", "_").replace("/", "_")

        filename, _ = QFileDialog.getSaveFileName(self, "Save PDF Report", suggested_name, "PDF Files (*.pdf)")
        
        if filename:
            try:
                doc = SimpleDocTemplate(filename, pagesize=letter)
                styles = getSampleStyleSheet()
                style_mono = styles['Code']
                
                formatted_text = report_text.replace('\\n', '<br/>')
                
                story = [Paragraph(formatted_text, style_mono)]
                
                doc.build(story)
                QMessageBox.information(self, "Success", f"Report successfully saved to:\\n{filename}")
            except Exception as e:
                QMessageBox.critical(self, "Error", f"Failed to save PDF: {str(e)}")

    
    def _write_excel_header(self, ws, row):
        """Writes the report header block to the worksheet."""
        header_lines = self.last_report_header_text.split('\\n')
        bold_font = Font(bold=True)
        for line in header_lines:
            if ":" in line:
                key, value = line.split(":", 1)
                ws.cell(row=row, column=1).value = key + ":"
                ws.cell(row=row, column=1).font = bold_font
                ws.cell(row=row, column=2).value = value.strip()
            else:
                ws.cell(row=row, column=1).value = line
            row += 1
        return row + 1

    def _auto_fit_excel_cols(self, ws):
        """Auto-fit all columns in the worksheet."""
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            ws.column_dimensions[column].width = adjusted_width

    def _download_excel(self) -> None:
        """Saves the generated report as a structured Excel file."""
        if not self.last_report_type:
            QMessageBox.warning(self, "Empty Report", "Cannot download an empty report. Please generate a report first.")
            return

        suggested_name = f"{self.last_report_org}_{self.last_report_type}_Report.xlsx".replace(" ", "_").replace("/", "_")
        filename, _ = QFileDialog.getSaveFileName(self, "Save Excel Report", suggested_name, "Excel Files (*.xlsx)")
        
        if not filename:
            return

        try:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Report"
            
            current_row = self._write_excel_header(ws, 1)
            
            bold_font = Font(bold=True)
            
            if self.last_report_type == "Officer List":
                ws.cell(row=current_row, column=1).value = "Name"
                ws.cell(row=current_row, column=1).font = bold_font
                ws.cell(row=current_row, column=2).value = "Position"
                ws.cell(row=current_row, column=2).font = bold_font
                current_row += 1
                
                if self.last_report_data:
                    for officer in self.last_report_data:
                        ws.cell(row=current_row, column=1).value = officer.get('name', 'N/A')
                        ws.cell(row=current_row, column=2).value = officer.get('position', 'N/A')
                        current_row += 1
                else:
                    ws.cell(row=current_row, column=1).value = "No officers found."

            elif self.last_report_type == "Summary":
                ws.cell(row=current_row, column=1).value = "Action"
                ws.cell(row=current_row, column=1).font = bold_font
                ws.cell(row=current_row, column=2).value = "Count"
                ws.cell(row=current_row, column=2).font = bold_font
                current_row += 1
                
                actions = {}
                for log in self.last_report_logs:
                    action_type = log.get('action', 'UNKNOWN')
                    actions[action_type] = actions.get(action_type, 0) + 1
                
                if not actions:
                    ws.cell(row=current_row, column=1).value = "No activities found."
                else:
                    for action, count in sorted(actions.items()):
                        ws.cell(row=current_row, column=1).value = action
                        ws.cell(row=current_row, column=2).value = count
                        current_row += 1
            
            else:
                headers = ["Timestamp", "Actor", "Action", "Subject", "Details"]
                for c, header in enumerate(headers, 1):
                    ws.cell(row=current_row, column=c).value = header
                    ws.cell(row=current_row, column=c).font = bold_font
                current_row += 1
                
                logs_to_write = self.last_report_logs
                if self.last_report_type == "Membership":
                    logs_to_write = [
                        log for log in self.last_report_logs
                        if log["action"] in ["ACCEPT_APPLICANT", "KICK_MEMBER", "EDIT_MEMBER"]
                    ]
                
                if not logs_to_write:
                    ws.cell(row=current_row, column=1).value = "No logs found for this report type."
                else:
                    for log in sorted(logs_to_write, key=lambda x: x['timestamp']):
                        try:
                            dt = datetime.datetime.fromisoformat(log['timestamp']).strftime('%Y-%m-%d %I:%M %p')
                        except (ValueError, TypeError):
                            dt = log.get('timestamp', 'N/A')
                            
                        ws.cell(row=current_row, column=1).value = dt
                        ws.cell(row=current_row, column=2).value = log.get('actor', 'N/A')
                        ws.cell(row=current_row, column=3).value = log.get('action', 'N/A')
                        ws.cell(row=current_row, column=4).value = log.get('subject', 'N/A')
                        ws.cell(row=current_row, column=5).value = log.get('details', 'N/A')
                        current_row += 1
            
            self._auto_fit_excel_cols(ws)
            wb.save(filename)
            QMessageBox.information(self, "Success", f"Report successfully saved to:\n{filename}")
            
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Failed to save Excel file: {str(e)}")

    def show_org_details(self, org_data: Dict) -> None:
        if org_data.get("is_archived", False):
            self.from_archived = True
        super().show_org_details(org_data)
        self.current_org = org_data
        self.ui.view_members_btn.setText("Manage Members")
        
        # Fetch pending applicants and active members for this organization
        self._fetch_applicants(org_data["id"])
        self._fetch_members(org_data["id"])
        
        # Load officers from fetched members data
        self._on_officer_history_changed(0)  # Load "Current Officers"
        
        # Load members view by default (this triggers the UI update)
        self.load_members()
        
        # Clean up existing admin buttons
        if self.archive_btn:
            self.ui.verticalLayout_10.removeWidget(self.archive_btn)
            self.archive_btn.deleteLater()
            self.archive_btn = None
        
        if hasattr(self, 'toggle_active_btn') and self.toggle_active_btn:
            self.ui.verticalLayout_10.removeWidget(self.toggle_active_btn)
            self.toggle_active_btn.deleteLater()
            self.toggle_active_btn = None
        
        # Only add admin buttons if org is not archived
        if not org_data.get("is_archived", False):
            edit_index = self.ui.verticalLayout_10.indexOf(self.edit_btn)
            insert_pos = edit_index + 1 if edit_index != -1 else -1
            
            # Add Toggle Active/Inactive button
            is_active = org_data.get("is_active", True)
            self.toggle_active_btn = QtWidgets.QPushButton("Deactivate" if is_active else "Activate")
            self.toggle_active_btn.setObjectName("toggle_active_btn")
            self.toggle_active_btn.clicked.connect(self._toggle_org_active)
            if is_active:
                self.toggle_active_btn.setStyleSheet("border-radius: 10px; background-color: #f0ad4e; color: white; border: 2px solid #f0ad4e")
            else:
                self.toggle_active_btn.setStyleSheet("border-radius: 10px; background-color: #5cb85c; color: white; border: 2px solid #5cb85c")
            
            if insert_pos != -1:
                spacer1 = QtWidgets.QSpacerItem(20, 10, QtWidgets.QSizePolicy.Policy.Minimum, QtWidgets.QSizePolicy.Policy.Fixed)
                self.ui.verticalLayout_10.insertItem(insert_pos, spacer1)
                self.ui.verticalLayout_10.insertWidget(insert_pos + 1, self.toggle_active_btn)
                insert_pos += 2
            else:
                self.ui.verticalLayout_10.addWidget(self.toggle_active_btn)
            
            # Add Archive button
            self.archive_btn = QtWidgets.QPushButton("Archive")
            self.archive_btn.setObjectName("archive_btn")
            self.archive_btn.clicked.connect(self._confirm_archive)
            self.archive_btn.setStyleSheet("border-radius: 10px; background-color: #EB5757; color: white; border: 2px solid #EB5757")
            
            if insert_pos != -1:
                spacer2 = QtWidgets.QSpacerItem(20, 10, QtWidgets.QSizePolicy.Policy.Minimum, QtWidgets.QSizePolicy.Policy.Fixed)
                self.ui.verticalLayout_10.insertItem(insert_pos, spacer2)
                self.ui.verticalLayout_10.insertWidget(insert_pos + 1, self.archive_btn)
            else:
                self.ui.verticalLayout_10.addWidget(self.archive_btn)
    
    def _toggle_org_active(self):
        """Toggle the organization's active/inactive status."""
        if not self.current_org:
            return
        
        org_id = self.current_org.get("id")
        is_active = self.current_org.get("is_active", True)
        action = "deactivate" if is_active else "activate"
        
        reply = QMessageBox.question(
            self,
            f"Confirm {action.capitalize()}",
            f"Are you sure you want to {action} '{self.current_org.get('name')}'?",
            QMessageBox.StandardButton.Yes | QMessageBox.StandardButton.No,
            QMessageBox.StandardButton.No
        )
        
        if reply == QMessageBox.StandardButton.Yes:
            from PyQt6.QtCore import QSettings
            from services.organization_api_service import OrganizationAPIService
            
            # Get admin user ID for logging
            settings = QSettings("CISC", "VirtualHub")
            admin_user_id = settings.value("profile_id")
            
            response = OrganizationAPIService.toggle_organization_active(org_id, admin_user_id)
            
            if response.get('success'):
                new_status = "activated" if response.get('is_active') else "deactivated"
                QMessageBox.information(self, "Success", f"Organization {new_status} successfully.")
                
                # Update current_org and refresh the view
                self.current_org['is_active'] = response.get('is_active')
                self.show_org_details(self.current_org)
                self.load_orgs()  # Refresh the org list
            else:
                QMessageBox.critical(self, "Error", f"Failed to {action} organization: {response.get('message')}")
            
    def _confirm_archive(self):
        """Show confirmation dialog and archive the organization via API."""
        if not self.current_org:
            return
        dialog = ArchiveConfirmDialog(self.current_org["name"], self)
        if dialog.exec() == QtWidgets.QDialog.DialogCode.Accepted:
            from PyQt6.QtCore import QSettings
            from services.organization_api_service import OrganizationAPIService
            
            org_id = self.current_org.get("id")
            
            # Get admin user ID for logging
            settings = QSettings("CISC", "VirtualHub")
            admin_user_id = settings.value("profile_id")
            
            response = OrganizationAPIService.archive_organization(org_id, admin_user_id)
            
            if response.get('success'):
                QMessageBox.information(self, "Success", "Organization archived successfully.")
                self._return_to_prev_page()
                self.load_orgs()
            else:
                QMessageBox.critical(self, "Error", f"Failed to archive organization: {response.get('message')}")

    def _fetch_applicants(self, org_id: int) -> None:
        """Fetch pending applicants for the organization from the API."""
        from services.organization_api_service import OrganizationAPIService
        
        api_response = OrganizationAPIService.get_organization_applicants(org_id)
        if api_response.get("success"):
            applicants = api_response.get("data", [])
            # Store applicants in current_org for load_applicants to use
            if self.current_org:
                self.current_org["applicants"] = applicants
        else:
            if self.current_org:
                self.current_org["applicants"] = []

    def _fetch_members(self, org_id: int) -> None:
        """Fetch active members for the organization from the API."""
        from services.organization_api_service import OrganizationAPIService
        
        api_response = OrganizationAPIService.get_organization_members(org_id)
        if api_response.get("success"):
            members_data = api_response.get("data", [])
            
            # Transform member data to match ViewMembers table format: [Name, Position, Status, Join Date]
            # But also store the full member dict for Edit functionality
            transformed_members = []
            self.members_dict = {}  # Store full member data by member_id
            
            for member in members_data:
                member_id = member.get('id')  # This is OrganizationMembers.id
                transformed = [
                    member.get('name', 'Unknown'),                                      # Name
                    member.get('position', 'Member'),                                   # Position (from backend)
                    'Active' if member.get('status') == 'active' else 'Inactive',      # Status
                    member.get('joined_at', '')[:10] if member.get('joined_at') else '' # Join Date (YYYY-MM-DD)
                ]
                transformed_members.append(transformed)
                # Store full member data keyed by member_id for Edit operations
                self.members_dict[member_id] = member
            
            # Store transformed members in current_org for load_members to use
            if self.current_org:
                self.current_org["members"] = transformed_members
        else:
            if self.current_org:
                self.current_org["members"] = []
                self.members_dict = {}
    
    def _update_member_in_list(self, member_id: int, updated_data: dict) -> None:
        """Update a single member in the members list without full refresh"""
        if not hasattr(self, 'members_dict') or not self.current_org:
            return
        
        try:
            # Update members_dict with new data
            self.members_dict[member_id] = updated_data
            
            # Find and update the member in current_org["members"]
            members = self.current_org.get("members", [])
            member_name = updated_data.get('name', 'Unknown')
            
            for idx, member in enumerate(members):
                if member[0] == member_name:  # Match by name
                    # Update the position (index 1)
                    new_position = updated_data.get('position', 'Member')
                    members[idx][1] = new_position
                    break
                
        except Exception:
            raise

    def _process_application_action(self, application_id: int, action: str, applicant_name: str):
        """Process accept or reject action via API - override to refresh members on accept"""
        from services.organization_api_service import OrganizationAPIService
        from PyQt6.QtWidgets import QMessageBox
        from PyQt6.QtCore import QSettings
        
        # Get the current user's profile_id for logging
        admin_user_id = None
        try:
            settings = QSettings("CISC", "MasterRoute")
            admin_user_id = settings.value("user_profile_id")
            if admin_user_id:
                admin_user_id = int(admin_user_id)
        except Exception:
            pass
        
        api_response = OrganizationAPIService.process_application(application_id, action, admin_user_id)
        
        if api_response.get("success"):
            action_text = "accepted" if action == "accept" else "rejected"
            QMessageBox.information(
                self,
                "Success",
                f"{applicant_name}'s application has been {action_text}."
            )
            # Refresh the applicants list and members list
            if self.current_org:
                self._fetch_applicants(self.current_org["id"])
                if action == "accept":
                    # Refresh members when accepting to show new member
                    self._fetch_members(self.current_org["id"])
                self.load_applicants(self._get_search_text())
        else:
            error_msg = api_response.get("error", f"Failed to {action} application")
            QMessageBox.critical(self, "Error", error_msg)
